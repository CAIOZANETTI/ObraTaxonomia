import re
import unicodedata
import os
from pathlib import Path
from typing import Optional, List, Dict, Tuple

def normalize_text(text):
    """
    Remove acentos, converte para minúsculas e remove caracteres especiais.
    Ex: 'Cimento Portland CP-II' -> 'cimento portland cp ii'
    """
    if not isinstance(text, str):
        return str(text).lower() if text is not None else ""
        
    # Unicode normalize (remove acentos)
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    text = text.lower()
    
    # Remove caracteres especiais (mantém apenas letras, números e espaços)
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    
    # Remove espaços duplos
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text

def normalize_unit(unit):
    """
    Normaliza unidades comuns.
    """
    if not isinstance(unit, str):
        return "un" # Default fallback
        
    u = normalize_text(unit)
    
    # Mapa simples de normalização (idealmente viria do yaml/unidades)
    # Aqui é um hardcode rápido para o script utils, mas o builder deve carregar do YAML.
    # Para simplificar este arquivo, vamos deixar genérico.
    return u


# ============================================================================
# XLSX to CSV Conversion Utilities - 8 Different Methods
# ============================================================================

def xlsx_to_csv_pandas(xlsx_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
    """
    Método 1: Pandas (Padrão de mercado para dados)
    Requer: pandas, openpyxl
    """
    try:
        import pandas as pd
        
        xlsx_file = pd.ExcelFile(xlsx_path)
        output_files = []
        
        for sheet_name in xlsx_file.sheet_names:
            df = pd.read_excel(xlsx_file, sheet_name=sheet_name)
            safe_name = re.sub(r'[^\w\s-]', '_', sheet_name)
            output_path = os.path.join(output_dir, f"{safe_name}.csv")
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            output_files.append(output_path)
        
        return True, f"Pandas: {len(output_files)} sheets converted", output_files
    except Exception as e:
        return False, f"Pandas failed: {str(e)}", []


def xlsx_to_csv_win32com(xlsx_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
    """
    Método 2: Win32com (Automação nativa usando Excel instalado)
    Requer: pywin32, Excel instalado no Windows
    """
    try:
        import win32com.client
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        abs_xlsx_path = os.path.abspath(xlsx_path)
        workbook = excel.Workbooks.Open(abs_xlsx_path)
        output_files = []
        
        for sheet in workbook.Sheets:
            sheet_name = sheet.Name
            safe_name = re.sub(r'[^\w\s-]', '_', sheet_name)
            output_path = os.path.abspath(os.path.join(output_dir, f"{safe_name}.csv"))
            
            # Salvar como CSV (formato 6 = CSV)
            sheet.SaveAs(output_path, FileFormat=6)
            output_files.append(output_path)
        
        workbook.Close(SaveChanges=False)
        excel.Quit()
        
        return True, f"Win32com: {len(output_files)} sheets converted", output_files
    except Exception as e:
        return False, f"Win32com failed: {str(e)}", []


def xlsx_to_csv_openpyxl(xlsx_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
    """
    Método 3: Openpyxl (Manipulação direta das células sem Pandas)
    Requer: openpyxl
    """
    try:
        from openpyxl import load_workbook
        import csv
        
        workbook = load_workbook(xlsx_path, data_only=True)
        output_files = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            safe_name = re.sub(r'[^\w\s-]', '_', sheet_name)
            output_path = os.path.join(output_dir, f"{safe_name}.csv")
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)
            
            output_files.append(output_path)
        
        return True, f"Openpyxl: {len(output_files)} sheets converted", output_files
    except Exception as e:
        return False, f"Openpyxl failed: {str(e)}", []


def xlsx_to_csv_polars(xlsx_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
    """
    Método 4: Polars (Performance ultra-rápida)
    Requer: polars, xlsx2csv ou openpyxl
    """
    try:
        import polars as pl
        from openpyxl import load_workbook
        
        workbook = load_workbook(xlsx_path, data_only=True)
        output_files = []
        
        for sheet_name in workbook.sheetnames:
            # Polars não lê XLSX diretamente, então usamos openpyxl para extrair dados
            sheet = workbook[sheet_name]
            data = list(sheet.iter_rows(values_only=True))
            
            if data:
                df = pl.DataFrame(data[1:], schema=data[0], orient='row')
                safe_name = re.sub(r'[^\w\s-]', '_', sheet_name)
                output_path = os.path.join(output_dir, f"{safe_name}.csv")
                df.write_csv(output_path)
                output_files.append(output_path)
        
        return True, f"Polars: {len(output_files)} sheets converted", output_files
    except Exception as e:
        return False, f"Polars failed: {str(e)}", []


def xlsx_to_csv_xlsx2csv(xlsx_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
    """
    Método 5: Xlsx2csv (Focado em baixo consumo de memória)
    Requer: xlsx2csv
    """
    try:
        from xlsx2csv import Xlsx2csv
        from openpyxl import load_workbook
        
        # Obter nomes das sheets
        workbook = load_workbook(xlsx_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        output_files = []
        
        for idx, sheet_name in enumerate(sheet_names):
            safe_name = re.sub(r'[^\w\s-]', '_', sheet_name)
            output_path = os.path.join(output_dir, f"{safe_name}.csv")
            
            # Converter sheet específica (índice baseado em 0)
            Xlsx2csv(xlsx_path, outputencoding="utf-8").convert(output_path, sheetid=idx+1)
            output_files.append(output_path)
        
        return True, f"Xlsx2csv: {len(output_files)} sheets converted", output_files
    except Exception as e:
        return False, f"Xlsx2csv failed: {str(e)}", []


def xlsx_to_csv_xlwings(xlsx_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
    """
    Método 6: Xlwings (Interface moderna para automação do Excel)
    Requer: xlwings, Excel instalado
    """
    try:
        import xlwings as xw
        
        app = xw.App(visible=False)
        workbook = app.books.open(xlsx_path)
        output_files = []
        
        for sheet in workbook.sheets:
            sheet_name = sheet.name
            safe_name = re.sub(r'[^\w\s-]', '_', sheet_name)
            output_path = os.path.join(output_dir, f"{safe_name}.csv")
            
            # Obter dados e salvar como CSV
            data = sheet.used_range.value
            
            import csv
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                if isinstance(data, list):
                    for row in data:
                        if isinstance(row, (list, tuple)):
                            writer.writerow(row)
                        else:
                            writer.writerow([row])
                else:
                    writer.writerow([data])
            
            output_files.append(output_path)
        
        workbook.close()
        app.quit()
        
        return True, f"Xlwings: {len(output_files)} sheets converted", output_files
    except Exception as e:
        return False, f"Xlwings failed: {str(e)}", []


def xlsx_to_csv_pyexcel(xlsx_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
    """
    Método 7: Pyexcel (API unificada e simples)
    Requer: pyexcel, pyexcel-xlsx
    """
    try:
        import pyexcel as pe
        
        book = pe.get_book(file_name=xlsx_path)
        output_files = []
        
        for sheet_name in book.sheet_names():
            sheet = book[sheet_name]
            safe_name = re.sub(r'[^\w\s-]', '_', sheet_name)
            output_path = os.path.join(output_dir, f"{safe_name}.csv")
            sheet.save_as(output_path)
            output_files.append(output_path)
        
        return True, f"Pyexcel: {len(output_files)} sheets converted", output_files
    except Exception as e:
        return False, f"Pyexcel failed: {str(e)}", []


def xlsx_to_csv_calamine(xlsx_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
    """
    Método 8: Python-Calamine (Leitura de alta velocidade baseada em Rust)
    Requer: python-calamine
    """
    try:
        from python_calamine import CalamineWorkbook
        import csv
        
        workbook = CalamineWorkbook.from_path(xlsx_path)
        output_files = []
        
        for sheet_name in workbook.sheet_names:
            sheet = workbook.get_sheet_by_name(sheet_name)
            safe_name = re.sub(r'[^\w\s-]', '_', sheet_name)
            output_path = os.path.join(output_dir, f"{safe_name}.csv")
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                for row in sheet.to_python():
                    writer.writerow(row)
            
            output_files.append(output_path)
        
        return True, f"Calamine: {len(output_files)} sheets converted", output_files
    except Exception as e:
        return False, f"Calamine failed: {str(e)}", []


def convert_xlsx_to_csv_all_methods(
    xlsx_path: str, 
    output_dir: Optional[str] = None,
    preferred_methods: Optional[List[str]] = None
) -> Dict[str, any]:
    """
    Tenta converter XLSX para CSV usando múltiplos métodos até um funcionar.
    
    Args:
        xlsx_path: Caminho para o arquivo XLSX
        output_dir: Diretório de saída (padrão: mesmo diretório do XLSX)
        preferred_methods: Lista de métodos preferidos na ordem (padrão: todos)
    
    Returns:
        Dict com:
            - success: bool
            - method: str (método que funcionou)
            - message: str
            - output_files: List[str]
            - attempts: List[Dict] (histórico de tentativas)
    """
    
    # Validar arquivo de entrada
    if not os.path.exists(xlsx_path):
        return {
            'success': False,
            'method': None,
            'message': f'Arquivo não encontrado: {xlsx_path}',
            'output_files': [],
            'attempts': []
        }
    
    # Definir diretório de saída
    if output_dir is None:
        output_dir = os.path.dirname(xlsx_path)
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Métodos disponíveis
    all_methods = {
        'pandas': xlsx_to_csv_pandas,
        'openpyxl': xlsx_to_csv_openpyxl,
        'xlsx2csv': xlsx_to_csv_xlsx2csv,
        'polars': xlsx_to_csv_polars,
        'win32com': xlsx_to_csv_win32com,
        'xlwings': xlsx_to_csv_xlwings,
        'pyexcel': xlsx_to_csv_pyexcel,
        'calamine': xlsx_to_csv_calamine,
    }
    
    # Definir ordem de tentativa
    if preferred_methods:
        methods_to_try = [(m, all_methods[m]) for m in preferred_methods if m in all_methods]
    else:
        # Ordem padrão: mais comuns primeiro
        methods_to_try = [
            ('pandas', xlsx_to_csv_pandas),
            ('openpyxl', xlsx_to_csv_openpyxl),
            ('xlsx2csv', xlsx_to_csv_xlsx2csv),
            ('polars', xlsx_to_csv_polars),
            ('win32com', xlsx_to_csv_win32com),
            ('xlwings', xlsx_to_csv_xlwings),
            ('pyexcel', xlsx_to_csv_pyexcel),
            ('calamine', xlsx_to_csv_calamine),
        ]
    
    attempts = []
    
    # Tentar cada método
    for method_name, method_func in methods_to_try:
        success, message, output_files = method_func(xlsx_path, output_dir)
        
        attempt = {
            'method': method_name,
            'success': success,
            'message': message
        }
        attempts.append(attempt)
        
        if success:
            return {
                'success': True,
                'method': method_name,
                'message': message,
                'output_files': output_files,
                'attempts': attempts
            }
    
    # Nenhum método funcionou
    return {
        'success': False,
        'method': None,
        'message': 'Todos os métodos falharam',
        'output_files': [],
        'attempts': attempts
    }

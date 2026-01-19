import pandas as pd
import io
import openpyxl
import xlrd
import os

# --- Copying logic from app/pages/1_Processar_Orcamento.py for testing ---
def list_sheets(file_bytes, ext):
    file_bytes.seek(0)
    sheets = []
    try:
        xls = pd.ExcelFile(file_bytes)
        sheets = xls.sheet_names
        return sheets
    except Exception:
        pass

    file_bytes.seek(0)
    if ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
            sheets = wb.sheetnames
            return sheets
        except Exception:
            pass

    file_bytes.seek(0)
    if ext == '.xls':
        try:
            wb = xlrd.open_workbook(file_contents=file_bytes.read())
            sheets = wb.sheet_names()
            return sheets
        except Exception:
            pass
            
    return []

def read_sheet(file_bytes, ext, sheet_name):
    file_bytes.seek(0)
    try:
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"
        df = pd.read_excel(file_bytes, sheet_name=sheet_name, header=None, engine=engine)
        return df, f"A) pd.read_excel (engine={engine})", None
    except Exception:
        pass

    file_bytes.seek(0)
    try:
        df_dict = pd.read_excel(file_bytes, sheet_name=None, header=None)
        if sheet_name in df_dict:
            return df_dict[sheet_name], "B) pd.read_excel (sheet_name=None)", None
    except Exception:
        pass

    file_bytes.seek(0)
    try:
        xls = pd.ExcelFile(file_bytes)
        df = xls.parse(sheet_name=sheet_name, header=None)
        return df, "C) pd.ExcelFile.parse", None
    except Exception:
        pass

    file_bytes.seek(0)
    if ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = list(ws.iter_rows(values_only=True))
                if data:
                    df = pd.DataFrame(data)
                else:
                    df = pd.DataFrame() 
                return df, "D) openpyxl (iter_rows)", None
        except Exception:
            pass

    file_bytes.seek(0)
    if ext == '.xls':
        try:
            wb = xlrd.open_workbook(file_contents=file_bytes.read())
            if sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                data = []
                for row_idx in range(sheet.nrows):
                    data.append(sheet.row_values(row_idx))
                if data:
                    df = pd.DataFrame(data)
                else:
                    df = pd.DataFrame()
                return df, "E) xlrd (row_values)", None
        except Exception as e:
            return None, None, str(e)

    return None, None, "Todos os métodos falharam."

# --- Testing ---
def run_tests():
    print("--- Testing XLSX ---")
    with open('tests/test_file.xlsx', 'rb') as f:
        file_bytes = io.BytesIO(f.read())
    
    sheets = list_sheets(file_bytes, '.xlsx')
    print(f"Sheets detected: {sheets}")
    assert 'Sheet1' in sheets
    assert 'Sheet2' in sheets
    
    df1, method, err = read_sheet(file_bytes, '.xlsx', 'Sheet1')
    print(f"Read Sheet1: {df1.shape}, Method: {method}, Error: {err}")
    assert df1 is not None
    
    print("\n--- Testing XLS (if exists) ---")
    if os.path.exists('tests/test_file.xls'):
        with open('tests/test_file.xls', 'rb') as f:
            file_bytes = io.BytesIO(f.read())
        
        sheets = list_sheets(file_bytes, '.xls')
        print(f"Sheets detected: {sheets}")
        assert 'Sheet1' in sheets
        
        df1, method, err = read_sheet(file_bytes, '.xls', 'Sheet1')
        print(f"Read Sheet1: {df1.shape}, Method: {method}, Error: {err}")
        assert df1 is not None
    else:
        print("Skipping XLS test (file not found)")

if __name__ == "__main__":
    run_tests()

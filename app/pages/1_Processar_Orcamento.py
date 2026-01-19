import streamlit as st
import pandas as pd
import io
import openpyxl
import xlrd
import traceback
from datetime import datetime

st.set_page_config(
    page_title="Upload e Conversão (Excel -> CSV)",
    page_icon="📂",
    layout="wide"
)

# --- 1. SESSION STATE INITIALIZATION ---
if 'excel_bytes' not in st.session_state:
    st.session_state['excel_bytes'] = None
if 'sheets_info' not in st.session_state:
    st.session_state['sheets_info'] = {} # {sheet_name: {rows, cols, status, read_method, error_msg, df_head}}
if 'df_all' not in st.session_state:
    st.session_state['df_all'] = None
if 'csv_all_bytes' not in st.session_state:
    st.session_state['csv_all_bytes'] = None
if 'df_resumo_abas' not in st.session_state:
    st.session_state['df_resumo_abas'] = None


# --- 2. HELPER FUNCTIONS ---

def list_sheets(file_bytes, ext):
    """
    Tenta listar as abas do arquivo usando diferentes estratégias.
    Retorna lista de nomes de abas.
    """
    file_bytes.seek(0)
    sheets = []
    
    # Estratégia 1: Pandas ExcelFile
    try:
        xls = pd.ExcelFile(file_bytes)
        sheets = xls.sheet_names
        return sheets
    except Exception:
        pass

    file_bytes.seek(0)
    # Estratégia 2: openpyxl (apenas xlsx)
    if ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
            sheets = wb.sheetnames
            return sheets
        except Exception:
            pass

    file_bytes.seek(0)
    # Estratégia 3: xlrd (apenas xls)
    if ext == '.xls':
        try:
            wb = xlrd.open_workbook(file_contents=file_bytes.read())
            sheets = wb.sheet_names()
            return sheets
        except Exception:
            pass
            
    return []

def read_sheet(file_bytes, ext, sheet_name):
    """
    Tenta ler uma aba específica usando cascata de 5 métodos.
    Retorna (df, read_method, error_msg)
    Se sucesso, error_msg é None. Se falha total, df é None.
    """
    file_bytes.seek(0)
    
    # --- Método A: pandas.read_excel (engine específica) ---
    try:
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"
        df = pd.read_excel(file_bytes, sheet_name=sheet_name, header=None, engine=engine)
        return df, f"A) pd.read_excel (engine={engine})", None
    except Exception:
        pass

    file_bytes.seek(0)
    # --- Método B: pandas.read_excel batch (sheet_name=None -> dict) ---
    # Nota: Isso é ineficiente para arquivos grandes se chamarmos para cada aba, 
    # mas mantendo a lógica isolada por aba como pedido, tentamos buscar no dict.
    try:
        # Tenta ler SÓ essa aba se possível, mas o método B pede sheet_name=None conceptualmente
        # porem para cumprir "ler todas as abas de uma vez" isso deveria ser feito fora.
        # Ajuste: Vamos tentar ler especificamente aqui como fallback de engine.
        df_dict = pd.read_excel(file_bytes, sheet_name=None, header=None)
        if sheet_name in df_dict:
            return df_dict[sheet_name], "B) pd.read_excel (sheet_name=None)", None
    except Exception:
        pass

    file_bytes.seek(0)
    # --- Método C: pd.ExcelFile + parse ---
    try:
        xls = pd.ExcelFile(file_bytes)
        df = xls.parse(sheet_name=sheet_name, header=None)
        return df, "C) pd.ExcelFile.parse", None
    except Exception:
        pass

    file_bytes.seek(0)
    # --- Método D: openpyxl (bruto) ---
    if ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = list(ws.iter_rows(values_only=True))
                if data:
                    df = pd.DataFrame(data)
                else:
                    df = pd.DataFrame() # Aba vazia
                return df, "D) openpyxl (iter_rows)", None
        except Exception:
            pass

    file_bytes.seek(0)
    # --- Método E: xlrd (bruto) ---
    if ext == '.xls':
        try:
            wb = xlrd.open_workbook(file_contents=file_bytes.read())
            if sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                data = []
                for row_idx in range(sheet.nrows):
                    data.append(sheet.row_values(row_idx))
                if data:
                    df = pd.DataFrame(data)
                else:
                    df = pd.DataFrame()
                return df, "E) xlrd (row_values)", None
        except Exception as e:
            return None, None, str(e)

    return None, None, "Todos os métodos falharam."


def build_summary(sheets_info):
    """
    Gera DataFrame de resumo: aba | linhas | colunas
    """
    data = []
    for s_name, info in sheets_info.items():
        rows = info.get('rows', 0)
        cols = info.get('cols', 0)
        # Se houve erro, rows/cols já devem vir como 0 do processamento
        data.append({
            'aba': s_name,
            'linhas': rows,
            'colunas': cols
        })
    return pd.DataFrame(data)

def consolidate_ok_sheets(sheets_info):
    """
    Consolida abas com status 'ok' e gera bytes do CSV.
    """
    dfs = []
    for s_name, info in sheets_info.items():
        if info['status'] == 'ok' and info.get('df_full') is not None:
            d = info['df_full'].copy()
            d['aba'] = s_name # Coluna obrigatória
            dfs.append(d)
    
    if dfs:
        # sort=False para manter ordem original e permitir schemas diferentes
        df_all = pd.concat(dfs, ignore_index=True, sort=False)
        return df_all
    else:
        return pd.DataFrame()


# --- 3. UI LAYOUT ---

st.title("📂 Upload e Conversão de Arquivo")
st.markdown("Carregue uma planilha **Excel (.xlsx ou .xls)** para gerar um **CSV Bruto Consolidado**.")

# Upload
uploaded_file = st.file_uploader("Selecione o arquivo Excel", type=["xlsx", "xls"])

if uploaded_file:
    # Verifica se arquivo mudou para resetar session state
    # (Streamlit recarrega script na interação, mas upload_file object muda ID se re-upado)
    # Aqui apenas processamos se ainda nao tivermos processado este arquivo específico ou se for novo
    
    # Para simplicidade: Processar sempre que houver upload ativo e botão acionado OU automático.
    # Vamos fazer automático para fluidez.
    
    file_bytes = io.BytesIO(uploaded_file.getvalue())
    filename = uploaded_file.name
    ext = ".xlsx" if filename.endswith(".xlsx") else ".xls"
    
    # Botão de processar (opcional, pode ser direto)
    # Fazendo direto para UX fluida
    
    # Verifica se já processamos (usando hash ou check simples se session state tem dados)
    # Por segurança, vamos reprocessar se st.session_state['excel_bytes'] for diferente
    # ou se o usuário acabou de dar upload.
    # Dado que file_uploader mantem o arquivo, vamos usar uma flag simples ou checar df_all
    
    # INÍCIO DO PROCESSAMENTO
    if st.session_state['df_all'] is None or st.session_state.get('last_filename') != filename:
        st.session_state['last_filename'] = filename
        st.session_state['excel_bytes'] = file_bytes
        
        with st.spinner("Lendo arquivo e varrendo abas..."):
            
            # 1. Listar abas
            sheet_names = list_sheets(file_bytes, ext)
            
            if not sheet_names:
                st.error("Não foi possível listar as abas deste arquivo. Verifique se é um Excel válido/desprotegido.")
            else:
                # 2. Varrer abas
                info_dict = {}
                
                prog_bar = st.progress(0)
                
                for idx, s_name in enumerate(sheet_names):
                    prog_bar.progress((idx) / len(sheet_names), text=f"Lendo aba: {s_name}")
                    
                    df, method, error = read_sheet(file_bytes, ext, s_name)
                    
                    s_info = {
                        'status': 'error',
                        'rows': 0,
                        'cols': 0,
                        'read_method': method,
                        'error_msg': error,
                        'df_head': None,
                        'df_full': None
                    }
                    
                    if df is not None:
                        rows, cols = df.shape
                        if rows == 0 and cols == 0:
                            s_info['status'] = 'empty'
                        else:
                            s_info['status'] = 'ok'
                            s_info['rows'] = rows
                            s_info['cols'] = cols
                            s_info['df_head'] = df.head(5)
                            s_info['df_full'] = df  # Guardamos df completo para consolidar (cuidado com memória em arquivos gigantes)
                    else:
                        s_info['status'] = 'error'
                        s_info['error_msg'] = error
                        
                    info_dict[s_name] = s_info
                
                prog_bar.empty()
                st.session_state['sheets_info'] = info_dict
                
                # 3. Gerar Resumo
                st.session_state['df_resumo_abas'] = build_summary(info_dict)
                
                # 4. Consolidar
                df_all = consolidate_ok_sheets(info_dict)
                st.session_state['df_all'] = df_all
                
                if not df_all.empty:
                    st.session_state['csv_all_bytes'] = df_all.to_csv(index=False).encode('utf-8')
                else:
                    st.session_state['csv_all_bytes'] = None

    
    # --- EXIBIÇÃO ---
    
    # 1. Expanders por Aba
    st.subheader("Seleção e Detalhamento por Aba")
    st.info("Desmarque as caixas para ignorar abas indesejadas na consolidação.")
    
    selected_sheets = []
    
    if st.session_state['sheets_info']:
        # Ensure selection state exists
        if 'sheet_selection' not in st.session_state:
            st.session_state['sheet_selection'] = {s: True for s in st.session_state['sheets_info']}
            
        for s_name, info in st.session_state['sheets_info'].items():
            status = info['status']
            icon = "✅" if status == 'ok' else "⚠️" if status == 'empty' else "❌"
            
            # Checkbox for selection (default True for OK sheets)
            # Use specific key to persist state
            is_ok = (status == 'ok')
            default_val = is_ok # Default checked only if OK
            
            # Update selection from session state logic if needed, but st.checkbox with key handles it
            
            c_check, c_expander = st.columns([0.05, 0.95])
            
            with c_check:
                # Disable if not OK? No, maybe user wants to inspect error but not select.
                # Only include in final DF if selected AND status is OK.
                is_selected = st.checkbox("Incluir", value=default_val, key=f"sel_{s_name}", label_visibility="collapsed")
            
            if is_selected and status == 'ok':
                selected_sheets.append(s_name)

            label = f"{icon} {s_name} (L:{info['rows']} x C:{info['cols']})"
            
            with c_expander:
                with st.expander(label):
                    c1, c2, c3 = st.columns(3)
                    c1.markdown(f"**Status:** {status.upper()}")
                    c2.markdown(f"**Linhas:** {info['rows']}")
                    c3.markdown(f"**Colunas:** {info['cols']}")
                    
                    st.caption(f"Método de leitura usado: {info.get('read_method')}")
                    
                    if status == 'error':
                        st.error(f"Erro: {info.get('error_msg')}")
                    elif status == 'ok':
                        st.dataframe(info['df_head'], use_container_width=True)
                    elif status == 'empty':
                        st.warning("Aba vazia.")

    # RE-CONSOLIDATE based on selection
    # We do this every rerun to ensure df_all matches selection
    final_dfs = []
    for s_name in selected_sheets:
        info = st.session_state['sheets_info'].get(s_name)
        if info and info['df_full'] is not None:
             d = info['df_full'].copy()
             d['aba'] = s_name
             final_dfs.append(d)
             
    if final_dfs:
        st.session_state['df_all'] = pd.concat(final_dfs, ignore_index=True, sort=False)
    else:
        st.session_state['df_all'] = pd.DataFrame()

    # 2. Tabela Resumo
    st.divider()
    st.subheader("Resumo do Arquivo (Consolidado)")
    if st.session_state['df_resumo_abas'] is not None:
        # Update summary with selection status?
        # Maybe just show stats of df_all
        pass

    # 3. Totais e Métricas
    if st.session_state['df_all'] is not None:
        total_rows = len(st.session_state['df_all'])
        total_cols = len(st.session_state['df_all'].columns)
        
        st.write(f"**Total de linhas no CSV consolidado:** {total_rows}")
        st.write(f"**Total de colunas no CSV consolidado:** {total_cols}")
        st.write(f"**Abas selecionadas:** {len(selected_sheets)} de {len(st.session_state['sheets_info'])}")

    # 4. Navigation (Download Removed)
    st.divider()
    
    if not st.session_state['df_all'].empty:
        if st.button("Avançar para Detecção de Cabeçalhos ➡️", type="primary"):
            st.switch_page("pages/2_Detectar_Cabecalhos.py")
    else:
        st.warning("Selecione pelo menos uma aba válida para prosseguir.")

else:
    # Reset state se remover arquivo
    st.session_state['excel_bytes'] = None
    st.session_state['df_all'] = None
    st.session_state['sheets_info'] = {}
